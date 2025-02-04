### IMPORTS ##
# math, data input, date
from cmath import nan
import pandas as pd
import numpy as np
from datetime import datetime, timedelta

# debug
import logging
from distutils.log import debug

# system
import os, re
from concurrent.futures import ThreadPoolExecutor, as_completed

from turtle import shape
from unicodedata import category
from xml.dom.expatbuilder import parseString

# Excel
from openpyxl import load_workbook
import xlrd
import zipfile
import xml.etree.ElementTree as ET

# GUI
from tkinter import filedialog

# plot
import matplotlib.colors as mcolors

# JVM
import jpype
import asposecells


class xlsxDecoder():
    '''Class to decode xlsx files.'''
    
    def __init__(self,xlsx_file='') -> None:
            '''Initialize the xlsxDecoder class.

            Attributes:
            xlsx_file -> str
                Path to xlsx file'''
            self.xlsx_file = xlsx_file

    def hex_to_rgb(self, hex_color):
        '''Convert a hexadecimal color (e.g. 'FF0000') to [R, G, B].
        
        Attributes:
        hex_color -> str
            Hexadecimal color code (with or without '#').'''
        hex_color = hex_color.lstrip('#')  # Entferne '#' falls vorhanden
        if len(hex_color) == 6:  # Nur RGB ohne Alpha
            r = int(hex_color[0:2], 16)
            g = int(hex_color[2:4], 16)
            b = int(hex_color[4:6], 16)
            return [r, g, b]
        elif len(hex_color) == 8:  # RGBA -> Ignoriere Alpha
            r = int(hex_color[2:4], 16)
            g = int(hex_color[4:6], 16)
            b = int(hex_color[6:8], 16)
            return [r, g, b]
        else:
            return [0, 0, 0]  # Standardwert für ungültige Farben

    def unzip_xlsx(self, sheet_nr):
        '''Unpack xlsx file and return the content of the sharedStrings.xml, styles.xml, theme.xml and the worksheet.
        
        Attributes:
        sheet_nr -> int
            Number of the sheet to be read.'''
        # Unpacking .xlsx-Datei
        with zipfile.ZipFile(self.xlsx_file, "r") as z:
            # Read sharedStrings.xml
            shared_strings = z.read("xl/sharedStrings.xml").decode("utf-8")
            # Read styles.xml
            styles = z.read("xl/styles.xml").decode("utf-8")
            # Read theme.xml
            themes = z.read("xl/theme/theme1.xml").decode("utf-8")
            # Read worksheet
            sheet_select = "sheet"+str(sheet_nr)+".xml"
            sheet = z.read("xl/worksheets/"+sheet_select).decode("utf-8") 
        return sheet, styles, themes, shared_strings

    def get_colors(self, styles_tree, namespace):
        '''Write the color palette from Excel.
        
        Attributes:
        styles_tree -> xml.etree.ElementTree
            XML tree of the styles.xml file.
        namespace -> dict
            Namespace for Excel files.'''
        colors = []
        colors_xml = styles_tree.find("main:colors", namespace)
        for indexed_colors_xml in colors_xml:
            for color in indexed_colors_xml.findall("main:rgbColor", namespace):
                if 'rgb' in color.attrib:  # Hexadecimal colors
                    rgb = color.attrib['rgb']
                    colors.append(self.hex_to_rgb(rgb))
                elif 'theme' in color.attrib:  # Theme-based colors
                    # Placeholder: Theme-Farben müssten aus einer separaten Definition extrahiert werden.
                    theme = color.attrib['theme']
                    colors.append(theme)
                elif 'indexed' in color.attrib:  # Indexed-Farben
                    indexed = color.attrib['indexed']
                    colors.append(indexed)

        return colors

    def get_themes(self, themes_tree, namespace):
        '''Write the theme colors out of Excel.
        
        Attributes:
        themes_tree -> xml.etree.ElementTree
            XML tree of the theme.xml file.
        namespace -> dict
            Namespace for Excel files.'''
        themes = []
        themes_xml = themes_tree.find("main:themeElements", namespace)
        for clrsheme in themes_xml:
            for indexed_themes_xml in clrsheme:
                theme_found = False
                for systheme in indexed_themes_xml.findall("main:sysClr", namespace):
                    hex_rgb = systheme.attrib['lastClr']
                    rgb = self.hex_to_rgb(hex_rgb)
                    if rgb == [255, 255, 255]:      # if systemcolor is white change it to black (valid color)
                        rgb = [0, 0, 0]
                    themes.append(rgb)
                    theme_found = True
                if not theme_found:
                    for systheme in indexed_themes_xml.findall("main:srgbClr", namespace):
                        rgb = systheme.attrib['val']
                        themes.append(self.hex_to_rgb(rgb))

        return themes

    def get_shared_strings(self, shared_tree, namespace, colors):
        '''Create a list of shared strings.
        
        Attributes:
        shared_tree -> xml.etree.ElementTree
            XML tree of the sharedStrings.xml file.
        namespace -> dict
            Namespace for Excel files.
        colors -> list
            List of colors in the Excel file.'''
        shared_list = []
        style_shared_list = []
        for si in shared_tree.findall("main:si", namespace):
            value = si.findall("main:t", namespace)
            if len(value) > 0:  # not nested values (in sharedStrings.xml)
                shared_list.append([value[0].text])
                style_props = si.findall("main:rPr", namespace)
                if style_props:
                    for style_prop in style_props:
                        if style_prop.find("main:color", namespace) is not None:
                            if style_prop.find("main:color", namespace).get("indexed") is not None:
                                color_style = colors[int(style_prop.find("main:color", namespace).get("indexed"))]
                            else:
                                color_style = None
                        else:
                            color_style = None

                        dict_style = {
                                "bold": style_prop.find("main:b", namespace) is not None,
                                "italic": style_prop.find("main:i", namespace) is not None,
                                "underline": style_prop.find("main:u", namespace) is not None,
                                "strike": style_prop.find("main:strike", namespace) is not None,
                                "font": style_prop.find("main:rFont", namespace).get("val") if style_prop.find("main:rFont", namespace) is not None else None,
                                "size": style_prop.find("main:sz", namespace).get("val") if style_prop.find("main:sz", namespace) is not None else None,
                                "color": color_style
                            }
                        style_shared_list.append([dict_style])
                else:
                    style_shared_list.append([np.nan])
            else: # nested values (in sharedStrings.xml)
                nested_xml = si.findall("main:r", namespace)
                nested_list = []
                style_nested_list = []
                for ne in nested_xml:
                    value_nested = ne.findall("main:t", namespace)
                    nested_list.append(value_nested[0].text)
                    style_props = ne.findall("main:rPr", namespace)
                    if style_props:
                        for style_prop in style_props:
                            if style_prop.find("main:color", namespace) is not None:
                                if style_prop.find("main:color", namespace).get("indexed") is not None:
                                    color_style = colors[int(style_prop.find("main:color", namespace).get("indexed"))]
                                else:
                                    color_style = None
                            else:
                                color_style = None

                            dict_style = {
                                "bold": style_prop.find("main:b", namespace) is not None,
                                "italic": style_prop.find("main:i", namespace) is not None,
                                "underline": style_prop.find("main:u", namespace) is not None,
                                "strike": style_prop.find("main:strike", namespace) is not None,
                                "font": style_prop.find("main:rFont", namespace).get("val") if style_prop.find("main:rFont", namespace) is not None else None,
                                "size": style_prop.find("main:sz", namespace).get("val") if style_prop.find("main:sz", namespace) is not None else None,
                                "color": color_style
                            }
                            style_nested_list.append(dict_style)
                    else:
                        style_nested_list.append(np.nan)
                shared_list.append(nested_list)
                style_shared_list.append(style_nested_list)

        return shared_list, style_shared_list

    def apply_tint(self, color, tint, black_theme):
        '''Apply the tint value to lighten or darken the color.
        
        Attributes:
        color -> int
            Color value (0-255).
        tint -> float
            Tint value (-1 to 1).
        black_theme -> bool
            True if the theme color is black.'''
        if tint > 0:    # brighter
            return int(color + (255 - color) * tint)
        else:
            if black_theme:     # more gray
                return int(255 * (1 + tint))
            else:       # darker
                return int(color * (1 + tint))

    def get_fonts(self, styles_tree, namespace, colors, themes):
        '''Create a list of fonts from styles.xml.
        
        Attributes:
        styles_tree -> xml.etree.ElementTree
            XML tree of the styles.xml file.
        namespace -> dict
            Namespace for Excel files.
        colors -> list
            List of colors in the Excel file.
        themes -> list
            List of themes in the Excel file.'''
        fonts = []

        for font in styles_tree.find("main:fonts", namespace):
            color_style = None
            if font.find("main:color", namespace) is not None:
                # Color indexed
                if font.find("main:color", namespace).get("indexed") is not None:
                    color_style = colors[int(font.find("main:color", namespace).get("indexed"))]
                else:
                    color_style = None
                # Theme indexed
                if font.find("main:color", namespace).get("theme") is not None and color_style is None:
                    theme_color = themes[int(font.find("main:color", namespace).get("theme"))]
                    tint_theme_color = font.find("main:color", namespace).get("tint")
                    if tint_theme_color:
                        if all(x == 0 for x in theme_color):
                            black_theme = True
                        else:
                            black_theme = False
                        color_style = [self.apply_tint(c, float(tint_theme_color), black_theme) for c in theme_color]
                    else:
                        color_style = theme_color
            else:
                color_style = None

            font_props = {
                "bold": font.find("main:b", namespace) is not None,
                "italic": font.find("main:i", namespace) is not None,
                "underline": font.find("main:u", namespace) is not None,
                "strike": font.find("main:strike", namespace) is not None,
                "font": font.find("main:rFont", namespace).get("val") if font.find("main:rFont", namespace) is not None else None,
                "size": font.find("main:sz", namespace).get("val") if font.find("main:sz", namespace) is not None else None,
                "color": color_style
            }
            fonts.append(font_props)
        
        return fonts


    def parse_styles_xml(self, numFmts, namespace):
        ''''Parse the numFmts XML file.
        
        Attributes:
        numFmts -> xml.etree.ElementTree
            XML tree of the numFmts file.
        namespace -> dict
            Namespace for Excel files.'''
        num_fmts = {}
        for numFmt in numFmts.findall("main:numFmt", namespace):
            numFmtId = int(numFmt.get("numFmtId"))
            formatCode = numFmt.get("formatCode")
            num_fmts[numFmtId] = formatCode

        return num_fmts

    def get_excel_format(self, value, numFmtId, num_fmts):
        '''Convert the Excel number format to a readable format.
        
        Attributes:
        value -> str
            Value to be converted.
        numFmtId -> int
            Number format ID.
        num_fmts -> dict
            Dictionary of number formats.'''
        
        excel_start_date = datetime(1899, 12, 30)  
        excel_numfmts = {
            0: "General",
            1: "0",
            2: "0.00",
            3: "#,##0",
            4: "#,##0.00",
            9: "0%",
            10: "0.00%",
            11: "0.00E+00",
            12: "# ?/?",
            13: "# ??/??",
            14: "DD/MM/YYYY",  
            15: "D-MMM-YY",
            16: "D-MMM",
            17: "MMM-YY",
            18: "h:mm AM/PM",
            19: "h:mm:ss AM/PM",
            20: "h:mm",
            21: "h:mm:ss",
            22: "DD/MM/YYYY h:mm",
            37: "#,##0 ;(#,##0)",  # Währung/Buchhaltung
            38: "#,##0 ;[Red](#,##0)",
            39: "#,##0.00;(#,##0.00)",
            40: "#,##0.00;[Red](#,##0.00)",
            45: "h:mm",
            46: "h:mm:ss",
            47: "h:mm:ss.0",
            48: "0.00E+00",
            49: "@",  # Text
        }

        numFmtId = int(numFmtId)

        if value is None:
            return None

        # Benutzerdefinierte Formate prüfen
        if numFmtId in num_fmts.keys():
            num_format = num_fmts[numFmtId]
        else:
            if numFmtId in excel_numfmts.keys():
                num_format = excel_numfmts[numFmtId]

        # Datumswerte umwandeln
        if numFmtId in range(14, 23):  # date
            if value.isdigit():  # numerical string
                return excel_start_date + timedelta(days=int(value))
            return value  

        # Prozentwerte
        if numFmtId in [9, 10]:  
            return f"{float(value) * 100:.2f}%" if isinstance(value, (int, float)) else value

        # Wissenschaftliche Notation
        if numFmtId == 11 or num_format == "0.00E+00":
            return f"{float(value):.2E}" if isinstance(value, (int, float)) else value

        # Brüche
        if numFmtId in [12, 13]:
            return float(value) if isinstance(value, (int, float)) else value

        # Währungen oder Zahlen mit Tausendertrennung
        if numFmtId in [3, 4, 37, 38, 39, 40]:
            return f"{float(value):,.2f}" if isinstance(value, (int, float)) else value

        # Text
        if numFmtId == 49 or num_format == "@":
            return str(value)

        # Standard-Zahlenformat
        return value

    def decoder(self, sheet_nr, logger) -> dict:
        '''Decode the xlsx file and return the content of the cells.
        
        Attributes:
        sheet_nr -> int
            Number of the sheet to be read.
        logger -> logging.Logger
            Logger object.'''
        # Namespace für Excel-Dateien
        namespace = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        theme_namespace = {"main": "http://schemas.openxmlformats.org/drawingml/2006/main"}

        logger.info('Funktion DECODER gestartet')

        sheet, styles, themes, shared_strings = self.unzip_xlsx(sheet_nr)
        # Parsen der XML-Dateien
        shared_tree = ET.fromstring(shared_strings)
        styles_tree = ET.fromstring(styles)
        themes_tree = ET.fromstring(themes)
        sheet_tree = ET.fromstring(sheet)

        # Get colors in excel file
        colors = self.get_colors(styles_tree, namespace)
        # print all colors in use as RGB values
        logger.debug('--- Print all color entities. ---')
        for id, color in enumerate(colors):
            logger.debug(f'Color ID {id}: {color}')

        # Get theme colors
        themes = self.get_themes(themes_tree, theme_namespace)
        # print all themes in use as RGB values
        logger.debug('--- Print all theme entities. ---')
        for id, theme in enumerate(themes):
            logger.debug(f'Color ID {id}: {theme}')

        # print xfs_cell ids
        logger.debug('--- Print all xfs_cell entities. ---')
        for id, xfs_cell in enumerate(styles_tree.find("main:cellXfs", namespace)):
            font_id = xfs_cell.attrib.get('fontId')
            numFmtId = xfs_cell.attrib.get('numFmtId')
            logger.debug(f'xfs_cell ID {id} has font ID {font_id} and number format ID {numFmtId}')

        # Get styles and values of cells containing numerous amount of different styles
        shared_list, style_shared_list = self.get_shared_strings(shared_tree, namespace, colors)

        # Get all possible fonts in excel file
        fonts = self.get_fonts(styles_tree, namespace, colors, themes)
        # Print fonts with their ids
        logger.debug('--- Print all font entities. ---')
        for id, font_prop in enumerate(fonts):
            logger.debug(f"Font ID {id}: {font_prop}")

        # Decode each cell. cell_data -> dict with each different formatted values as entries of a list with its style in a different list
        cell_data = {}

        # print styles of shared strings
        logger.debug('--- Print all style entities (if exist) of all shared strings. ---')
        for id, shared_str in enumerate(style_shared_list):
            logger.debug(f'Shared Style ID {id} for shared value {shared_list[id]} has shared style: {shared_str}')

        standard_color = [0, 0, 0]

        for row in sheet_tree.find("main:sheetData", namespace):
            for cell in row:
                cell_ref = cell.attrib.get("r")  # Zellreferenz (z. B. "A1")
                cell_type = cell.attrib.get("t")  # Zelltyp
                style_index = cell.attrib.get("s")  # Stilindex

                # default values
                cell_value = None
                cell_style = {}

                # 1. Identify cell value and style
                if cell_type == "s":  # Shared String
                    value_index = int(cell.find("main:v", namespace).text)
                    cell_value = shared_list[value_index]
                    cell_style = style_shared_list[value_index]
                else:
                    value_element = cell.find("main:v", namespace)
                    if value_element is not None:
                        cell_value = [value_element.text]

                # 2. Add style to each cell value
                if style_index is not None:
                    style_index = int(style_index)
                    font_id = None
                    cellXfs = styles_tree.find("main:cellXfs", namespace)
                    # 2.1. Check if <cellXfs> is defined
                    try:
                        target_style = cellXfs.findall("main:xf", namespace)[style_index]
                        font_id = target_style.attrib.get('fontId') 
                        numFmtId = target_style.attrib.get('numFmtId')  
                        if font_id is not None:
                            font_id = int(font_id)
                        if numFmtId is not None:
                            numFmts = styles_tree.find("main:numFmts", namespace)
                            num_fmt = self.parse_styles_xml(numFmts, namespace)
                            if cell_value is not None:
                                for id, value in enumerate(cell_value):
                                    cell_value[id] = self.get_excel_format(value, numFmtId, num_fmt)
                    except (IndexError, ValueError):
                        font_id = 0 # font-ID ungültig

                    # 2.2. Fallback
                    if fonts[font_id]["color"] is None or fonts[font_id]["color"] is np.nan:
                        fonts[font_id]["color"] = standard_color
                        
                    # 2.3. Check of dynamic styles in <dxfs>
                    dxfs = styles_tree.find("main:dxfs", namespace)
                    if dxfs is not None and font_id is None:
                        dynamic_style = dxfs.findall("main:dxf", namespace)
                        font_id = dynamic_style.attrib.get('fontId')
                        if font_id is not None:
                            font_id = int(font_id)

                    # 2.4. Check if font_id exists and if color in nested_xml is None
                    standard_style = fonts[font_id] if font_id < len(fonts) and font_id is not None else {}     # style of the whole cell
                    cell_style_nested = []
                    if cell_type == "s":
                        if fonts[font_id]["color"] is not None:
                            if isinstance(shared_list[value_index], list) and len(shared_list[value_index]) > 1:  # nested styles
                                for embedded_style in style_shared_list[value_index]:
                                    if embedded_style is np.nan:
                                        cell_style_nested.append(standard_style)    # use style of the whole cell
                                    elif embedded_style["color"] is None:
                                        embedded_style_color = embedded_style
                                        embedded_style_color["color"] = fonts[font_id]["color"]
                                        cell_style_nested.append(embedded_style_color)
                                    else:
                                        cell_style_nested.append(embedded_style)
                                cell_style = cell_style_nested
                            else:
                                if style_shared_list[value_index][0] is np.nan:
                                    cell_style = [standard_style]
                                else:
                                    cell_style = [style_shared_list[value_index][0]]
                    else:
                        cell_style = [standard_style]


                # 3. Write data in dict
                cell_data[cell_ref] = {
                    "value": cell_value,
                    "type": cell_type,
                    "style": cell_style,
                }

        return cell_data

class xlsxParser():
    '''Class to parse the contents of each cell of a excel file.'''

    # def __init__(self, path: str) -> None:
    #     return self.get_path_info(path)
    def __init__(self, color_valid = [], metadata_vertical = True, trigger_metadata = '', double_metadata = False, 
    trigger_double_metadata = '', main_category = True, main_category_pos = -1, sub_category= True, exclude_in_subcategory = [], recog_format=True) -> None:
        '''Initialize the xlsxParser class.
        
        Attributes:
        color_valid -> list
            List of valid colors.
        metadata_vertical -> bool
            True if metadata is vertical.
        trigger_metadata -> str
            Trigger metadata.
        double_metadata -> bool
            True if metadata is double.
        trigger_double_metadata -> str
            Trigger for double metadata.
        main_category -> bool
            True if main category is present.
        main_category_pos -> int
            Position of main category.
        sub_category -> bool
            True if sub category is present.
        exclude_in_subcategory -> list
            List of excluded subcategories.
        recog_format -> bool
            True if format is recognized.'''
        self.vehicle_number = []
        self.vehicle_series = []
        self.vehicle_plattform = []
        self.vehicle_engine_gearbox = []
        self.vehicle_weight = []
        self.vehicle_test = []
        self.vehicle_plate = []
        self.endurance_number = []

        self.brake_type = []
        self.brake_size = []
        self.brake_cylinder_diam = []
        self.brake_pad_material = []

        self.vds_number = []
        self.record_city = []
        self.record_date = []
        self.path_info_str = ""

        colors = [col.lower() if isinstance(col, str) else col for col in color_valid]
        self.color_valid = colors       # can be a letter (k: black, r: red, g: green) or a word (black, red, green) or a RGB als tuple
        self.metadata_vertical = metadata_vertical
        self.trigger_metadata = trigger_metadata
        self.double_metadata = double_metadata
        self.trigger_double_metadata = trigger_double_metadata
        self.sub_category = sub_category      # BE CAREFUL: if metadata of subcategories are of equal named, CONCAT/INGEST TO SQL will not work
        self.exclude_in_subcategory = exclude_in_subcategory    # metadata included by self.trigger_double_metadata but should not have a subcategory
        self.main_category = main_category      # BE CAREFUL: if metadata of non-subcategories are of equal named, CONCAT/INGEST TO SQL will not work
        self.main_category_pos = main_category_pos     # location relative to metadata
        self.recog_format = recog_format    # if TRUE: metadata will be removed not fulfilling certain condittions
        
    def setup_logger(self, log_file):
        '''Set up the logger.
        
        Attributes:
        log_file -> str
            Path to the log file.'''
        # Logger konfigurieren
        logger = logging.getLogger('FORMAT-LOG')
        logger.setLevel(logging.DEBUG)
        
        # Datei-Handler hinzufügen
        file_handler = logging.FileHandler(log_file)
        file_handler.setLevel(logging.DEBUG)
        
        # Format für Logs festlegen
        formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
        file_handler.setFormatter(formatter)
        
        # Handler zum Logger hinzufügen
        logger.addHandler(file_handler)
        
        return logger    


    def get_pattern(self) -> str:
        '''Get the pattern for the endurance number.'''

        pattern = r'((?P<endurance_number>DL\d{4}[SBLK]?)\s?[_]?[\\]?)?'

        return pattern

    def get_trigger_metadata(self, df, matches):
        '''Get row and colum of metadata trigger of metadata identifying the header of the data.
        
        Attributes:
        df -> pd.DataFrame
            Dataframe of the xlsx file.
        matches -> list
            List of matches.'''
        # Ausgabe des Index und der Spalte des trigger of metadata
        for idx, col_trigger_metadata in matches:
            # print(f"Fahrzeuge gefunden in Zeile: {idx+1}, Spalte: {col_trigger_metadata}")
            break

        row_values_trigger_metadata = df.loc[idx].values
        row_trigger_metadata = idx

        trigger_metadata = {}
        trigger_metadata_bool = False
        iter_nr = 1
        # iterate through row of trigger (e. g. vehicle numbers)
        for idx, nr in enumerate(row_values_trigger_metadata):
            if nr == self.trigger_metadata:
                trigger_metadata_bool = True  # avoiding initiliazing key for non-vehicle-numbers
                continue
            
            if nr is not None and trigger_metadata_bool:
                nr_found = str(nr)
                if nr_found in trigger_metadata.keys():   # == 1, if vehicle already exists in list but is not the same vehicle according to the column index
                    nr_found += '_v'+str(iter_nr)
                    while nr_found in trigger_metadata.keys():
                        iter_nr += 1
                        nr_found += '_v'+str(iter_nr)
                    trigger_metadata[nr_found] = []
                    iter_nr = 1
                else:
                    trigger_metadata[nr_found] = []
                trigger_metadata[nr_found].append(idx)
            
            
        return row_trigger_metadata, col_trigger_metadata, trigger_metadata

    def get_metadata_naming(self, df, row_lfd_nr, col_lfd_nr):
        '''Get the metadata names.
        
        Attributes:
        df -> pd.DataFrame
            Dataframe of the xlsx file.
        row_lfd_nr -> int
            Row number of the metadata.
        col_lfd_nr -> int
            Column number of the metadata.'''
        # return self.remove_trailing_nan(df[col_lfd_nr].iloc[row_lfd_nr+1:].to_numpy())
        return df[col_lfd_nr].iloc[row_lfd_nr+1:].to_numpy()

    def remove_trailing_nan(self, arr):
        '''Remove trailing NaN values from an array.
        
        Attributes:
        arr -> np.array
            Array to be processed.'''
        # Finde Indizes der gültigen Werte (alles außer nan)
        valid_index = np.argwhere(arr == arr)
        if valid_index.size > 0:  # Überprüfen, ob es gültige Werte gibt
            last_valid = valid_index[-1][0]
            if last_valid == len(arr):
                return arr
            else:
                return arr[:last_valid + 1]
        else:
            return np.array([])  # Leeres Array zurückgeben, falls nur "NaN"

    def get_duplicate_columns(self, df):
        '''Get the duplicate columns.
        
        Attributes:
        df -> pd.DataFrame
            Dataframe of the xlsx file.'''
        duplicates = df.columns[df.columns.duplicated()].unique()
        duplicates_positions = {col: [i for i, x in enumerate(df.columns) if x == col] for col in duplicates}
        return duplicates_positions

    def reset_duplicate_columns(self, df, category_names):
        '''IDEA: There exist many metadata with the same name. However, those metadata correspond to a certain subcategory. 
        To find those subcategory names a trigger is used. The trigger is a metadatum itself. The following metadatum to 
        this trigger is the repective metadatum.

        self.trigger_double_metadata : Trigger for subcategory

        Attributes:
        df -> pd.DataFrame
            Dataframe of the xlsx file.
        category_names -> dict
            Dictionary of category names.'''
        # get all ducplicate columns
        duplicates_positions = self.get_duplicate_columns(df)
        # print(f"Doppelte Spaltennamen und deren Positionen: {duplicates_positions}")
        not_notice = []
        for col, positions in duplicates_positions.items():
            for excluded_metadata_in_subcategory in self.exclude_in_subcategory:
                if str(excluded_metadata_in_subcategory).lower() in str(col).lower():
                    for pos in positions:
                        not_notice.append(pos)
        # print(f'not_notice: {not_notice}')

        # find the names of the subcategories being the next fields after the "note" field and rename
        duplicates_lower = [str(key).lower() for key in duplicates_positions.keys()]
        if any(self.trigger_double_metadata in s for s in duplicates_lower):
            
            for key, values in duplicates_positions.items():
                duplicates_positions[key] = np.flip(values)
                # print('Flipped Indizes für {1} der Duplikate: {0}'.format(duplicates_positions[key],key))

            last_note_excluded = False      # False: Exclude last trigger since there is no subcategory after this last trigger
            last_note_position = 0
            for duplicates_value, duplicates_position in duplicates_positions.items():  # if mulitple note fields exist
                if self.trigger_double_metadata in str(duplicates_value).lower():       # go through all positions where the self.trigger_double_metadata is shown!

                    iter_unnamed = 1    # Iterator for unnamend columns

                    # print('Für {0}: Flipped Indizes der Duplikate: {1}'.format(duplicates_value,duplicates_position))
                    for pos in duplicates_position:
                        row_to_read = pos + 1       # read metadatum after trigger field
                        if row_to_read in not_notice:       # add +1 since a metadatum is after the trigger which is not a subcategory nor a specific metadatum must be added to a subcategory
                            df.columns.values[row_to_read] = df.columns[row_to_read]
                            row_to_read += 1
                        for _, idx_cat in category_names.items():     # main categories
                            if row_to_read in idx_cat:      # save first occurence of the main category for the current subcategory
                                last_cat_idx = idx_cat[0]
                                continue
                        if last_note_excluded:
                            subcategory = df.columns[row_to_read]       # read subcategory name after trigger
                            # print('Subkategorie: {0}'.format(subcategory))
                            for idx, subcategory_col in enumerate(df.columns[row_to_read:last_note_position+1]):      # last_note_position = pos of the last iteration; +2 to include the trigger in the given subcategory
                                if idx > 0:      
                                    add_subcategory = row_to_read + idx
                                    if pd.isna(subcategory_col):
                                        subcategory_col = self.get_nan_columns(iter_unnamed)
                                        iter_unnamed += 1
                                    # print('In Zeile {0} füge Metadatum {1}${2} hinzu'.format(add_subcategory,subcategory,subcategory_col), flush=True)
                                    df.columns.values[add_subcategory] = subcategory + '$' + subcategory_col
                        last_note_excluded = True
                        last_note_position = pos
                    if duplicates_position[-1] > last_cat_idx:          # if the first subcategory is not after the given trigger
                        subcategory = df.columns[last_cat_idx]
                        # print('Subkategorie: {0}'.format(subcategory))
                        for idx, subcategory_col in enumerate(df.columns[last_cat_idx:last_note_position+1]):       # +2 to include the trigger in the given subcategory
                            if idx > 0:
                                add_subcategory = last_cat_idx + idx
                                # print('In Zeile {0} füge Metadatum {1} hinzu'.format(add_subcategory,subcategory_col))
                                if pd.isna(subcategory_col):
                                        subcategory_col = self.get_nan_columns(iter_unnamed)
                                        iter_unnamed += 1
                                df.columns.values[add_subcategory] = subcategory + '$' + subcategory_col
                                # print('In Zeile {0} füge Metadatum {1}${2} hinzu'.format(add_subcategory,subcategory,subcategory_col), flush=True)
        else:
            print('Keine Bemerkungsfelder im Template.')

        return df, iter_unnamed

    def get_category_names(self, df, index_spalte, row_lfd_nr):
        '''Get the category names.
        
        Attributes:
        df -> pd.DataFrame
            Dataframe of the xlsx file.
        index_spalte -> int
            Column index.
        row_lfd_nr -> int
            Row number.'''
        cat_names = {}
        if index_spalte > 0:
            cat_name = ''
            for idx, cat in enumerate(df[df.columns[index_spalte - 1]]):
                if pd.notna(cat):
                    cat_name = cat
                if cat_name:
                    if cat_name not in cat_names.keys():
                        cat_names[cat_name] = []
                    cat_names[cat_name].append(idx-row_lfd_nr)
                    if len(cat_names[cat_name]) > 1:
                        self.double_metadata = True
        # print(f'Main category names: {cat_names}')

        return cat_names

    def set_nan_columns(self, iter):
        '''Set NaN columns.
        
        Attributes:
        iter -> int
            Iterator.'''
        return 'Unnamed: '+str(iter)

    # Funktion zum Finden des Strings
    def find_string_in_dataframe(self, df):
        '''Find the string in the dataframe.
        
        Attributes:
        df -> pd.DataFrame
            Dataframe of the xlsx file.'''
        result = []
        for col in df.columns:
            # Überprüfen, ob der String in der Spalte vorkommt
            matching_indices = df[df[col] == self.trigger_metadata].index.tolist()
            for idx in matching_indices:
                result.append((idx, col))
        return result


    def data_cleaning(self, df, pat):
        '''Clean the data. Use this function to start the excel cleaning after reading it with function read_data. After applying this function
        only excel cell entries with valid format will stay in the pandas dataframe.
        
        Attributes:
        df -> pd.DataFrame
            Dataframe of the xlsx file.
        pat -> str
            Path of the xlsx file.'''
        matches = self.find_string_in_dataframe(df)
        df_vehicle, df_pattern = self.get_single_columns(df, matches)
        df_vehicle['Pfad'] = pat

        return df_vehicle, df_pattern

    def conditional_join(self, row):
        '''Conditional join with | for cell of combined columns.
        
        Attributes:
        row -> pd.Series
            Series of the row.'''
        result = []
        for val in row:
            if pd.notna(val):
                result.append(str(val))
        return '|'.join(result)

    def get_single_columns(self, df, matches):
        '''Get the single columns and put Excel data in dataframe with header row-wise sorted.
        
        Attributes:
        df -> pd.DataFrame
            Dataframe of the xlsx file.
        matches -> list
            List of matches.'''
       # find IDs based on trigger_metadata variable
        row_trigger_metadata, col_trigger_metadata, trigger_metadata_dict = self.get_trigger_metadata(df, matches)

        # combine columns of same lfd. Nr.
        df_vehicle = pd.DataFrame()
        if not self.metadata_vertical:
            df = df.T
        for keys, values in trigger_metadata_dict.items():
            column_name_target = keys
            column_name = df.columns[values]
            df_vehicle[column_name_target] = df[column_name].apply(self.conditional_join, axis=1)
        
        vehicle_nr = df_vehicle.iloc[row_trigger_metadata]
        df_vehicle = df_vehicle.replace(["", " ", None], np.nan)
        df_vehicle.columns = vehicle_nr
        df_vehicle = df_vehicle.iloc[row_trigger_metadata+1:].reset_index(drop=True)

        # get metadata naming
        index_name = self.get_metadata_naming(df, row_trigger_metadata, col_trigger_metadata)
        # print(f'index_name: {index_name}')
        index_name = pd.Series(index_name)
        df_vehicle = df_vehicle.reindex(index_name.index, fill_value=nan)
        df_vehicle[self.trigger_metadata] = index_name
        df_vehicle.set_index(self.trigger_metadata, inplace=True)

        # get pattern in dataframe
        df_pattern = self.get_dataframe_info(df_vehicle)

        # transpose dataframe: col = metadata, index = vehicles
        col_names = df_vehicle.columns
        df_vehicle = df_vehicle.transpose()
        df_vehicle.index.name = self.trigger_metadata
        df_vehicle.index = col_names
        df_vehicle = df_vehicle.reset_index()

        # get and set duplicate columns
        index_spalte = df.columns.get_loc(col_trigger_metadata)
        if self.sub_category:
            category_names = self.get_category_names(df, index_spalte, row_trigger_metadata)
            if self.double_metadata:
                df_vehicle, iter_unnamed = self.reset_duplicate_columns(df_vehicle, category_names)

        # get categories of metadata naming
        if self.main_category:
            if index_spalte > 0:
                cat_name = ''
                for idx, cat in enumerate(df[df.columns[index_spalte + self.main_category_pos]]):
                    if pd.notna(cat):
                        cat_name = cat
                    if cat_name:
                        # print('Kategorie: {0}'.format(cat_name))
                        if idx == 0:
                            # print(cat_name + '$' + str(df_vehicle.index.name))
                            df_vehicle.index.name = cat_name + '$' + str(df_vehicle.index.name)
                        else:
                            # print(cat_name + '$' + str(df_vehicle.columns.values[idx-row_trigger_metadata]))
                            colname = df_vehicle.columns.values[idx-row_trigger_metadata]
                            if pd.isna(df_vehicle.columns.values[idx-row_trigger_metadata]):
                                    colname = self.set_nan_columns(iter_unnamed)
                                    iter_unnamed += 1
                            df_vehicle.columns.values[idx-row_trigger_metadata]  = cat_name + '$' + str(colname)
                            # df hat mehr Zeilen an Metadaten als df_vehicle, da df_vehicle die Spalten nicht bis zum Ende besetzt 

        df_vehicle = df_vehicle.dropna(how="all")    # drop all rows showing only NaN values
        
        return df_vehicle, df_pattern
    
    def fix_hyphenated_words(self, text):
        '''Fix hyphenated words in a text removing return an double space.
        
        Attributes:
        text -> str
            Text to be fixed.'''
        if isinstance(text, str):  # Nur mit Strings arbeiten
            # Überprüfen, ob ein Zeilenumbruch und ein Bindestrich vorhanden sind
            if "\n" in text:
                # Text an Zeilenumbrüchen teilen
                parts = text.split("\n")
                # Überprüfen, ob das letzte Wort mit einem Bindestrich endet
                if parts[0].endswith("-"):
                    # Verbinde das Wort und entferne den Bindestrich
                    fixed_text = parts[0][:-1] + ''.join(parts[1:])
                    return fixed_text
                else:
                    fixed_text = parts[0] + ' ' + ' '.join(parts[1:])
                    return fixed_text
        return text

    def read_data(self, data_path):
        '''Read the data from the xlsx file. Start this function at first to read the data from the xlsx file.
        
        Attributes:
        data_path -> str
            Path to the xlsx file'''
        print('-------- Read file: '+data_path+' --------')
        excel_file = pd.ExcelFile(data_path)
        sheet_names = [(id, sheet) for id, sheet in enumerate(excel_file.sheet_names) if sheet.startswith('Fzg')]

        dfs = {sheet: excel_file.parse(sheet, header=None) for id, sheet in sheet_names}
        matches = {sheet: self.find_string_in_dataframe(dfs[sheet]) for id, sheet in sheet_names}

        if self.recog_format: # format = true?
            dfs = self.get_data_formatting(sheet_names, excel_file, data_path, matches)
        else:
            dfs = {sheet: excel_file.parse(sheet, header=None) for id, sheet in sheet_names}

        return dfs

    def get_color_palette(self, file_path):
        '''Get the color palette of the xlsx file.
        
        Attributes:
        file_path -> str
            Path to the xlsx file.'''
        workbook = xlrd.open_workbook(file_path, formatting_info=True)
        palette = workbook.colour_map
        return palette

    def compare_font_color(self, color) -> bool:
        '''Return TRUE, if a given color is required to stay in dataframe. Type specific RGB-values as a tuple or the following colors:
        'b'/'blue' für Blau
        'g'/'green' für Grün
        'r'/'red' für Rot
        'c'/'cyan' für Cyan
        'm'/'magenta' für Magenta
        'y'/'yellow' für Gelb
        'k'/'black' für Schwarz
        
        Cells are read under all circumstances if their font is standard colored in Excel.
        
        Attributes:
        color -> tuple
            RGB-values of the color.'''
        if not color: # default color (mostly black)
            return True
        elif color[0] == 0 and color[1] == 0 and color[2] == 0: # black
            if 'k' in self.color_valid or 'black' in self.color_valid:
                return True
            else:
                return False
        elif color[0] > 0 and color[1] == 0 and color[2] == 0: # red
            if 'r' in self.color_valid or 'red' in self.color_valid:
                return True
            else:
                return False
        elif color[0] == 0 and color[1] > 0 and color[2] == 0: # green
            if 'g' in self.color_valid or 'green' in self.color_valid:
                return True
            else:
                return False
        elif color[0] == 0 and color[1] == 0 and color[2] > 0: # blue
            if 'b' in self.color_valid or 'blue' in self.color_valid:
                return True
            else:
                return False
        elif color[0] == 0 and color[1] == 255 and color[2] == 255: # cyan
            if 'c' in self.color_valid or 'cyan' in self.color_valid:
                return True
            else:
                return False
        elif color[0] == 255 and color[1] == 0 and color[2] == 255: # magenta
            if 'm' in self.color_valid or 'magenta' in self.color_valid:
                return True
            else:
                return False
        elif color[0] == 255 and color[1] == 255 and color[2] == 0: # yellow
            if 'y' in self.color_valid or 'yellow' in self.color_valid:
                return True
            else:
                return False
        else:
            for item in self.color_valid:
                if isinstance(item, tuple) and item == (color[0], color[1], color[2]):
                    return True
            return False

    def convert_xls2xlsx(self,xls_file_path,target_folder,extension='converted.xlsx'):
        '''Convert .xls file to .xlsx file.
        
        Attributes:
        xls_file_path -> str
            Path to the .xls file.
        target_folder -> str
            Path to the target folder.
        extension -> str
            Extension of the converted file.'''
        jpype.startJVM() 
        from asposecells.api import Workbook

        print(f'xls_file_path: {xls_file_path}')
        # load .xls file in path
        workbook = Workbook(xls_file_path)

        filename_with_extension = os.path.basename(xls_file_path)
        filename, _ = os.path.splitext(filename_with_extension)
        filename = filename+'_'+extension
        xlsx_file_path = os.path.join(target_folder, filename)
        print(f'xlsx_file_path: {xlsx_file_path}')

        # save .xls as .xlsx file
        workbook.save(xlsx_file_path)
        # shut down JVM
        jpype.shutdownJVM()

        return xlsx_file_path

    def get_data_formatting(self, sheet_names, excel_file, data_path, matches):
        '''Get the data formatting of the xlsx file.
        
        Attributes:
        sheet_names -> list
            List of sheet names.
        excel_file -> pd.ExcelFile
            Excel file.
        data_path -> str
            Path to the xlsx file.
        matches -> list
            List of matches.'''
        user = os.environ['USERPROFILE']
        if not os.path.exists(user+r'\.bigbrems'):
            os.makedirs(user+r'\.bigbrems')
        if not os.path.exists(user+r'\.bigbrems\logs'):
            os.makedirs(user+r'\.bigbrems\logs')
        if not os.path.exists(user+r'\.bigbrems\logs\endu_nvh'):
            os.makedirs(user+r'\.bigbrems\logs\endu_nvh')
        folder = user+r'\.bigbrems\logs\endu_nvh'
        filename_with_extension = os.path.basename(data_path)
        filename, ext = os.path.splitext(filename_with_extension)

        subfolder = os.path.basename(os.path.dirname(data_path))
        
        if not os.path.exists(folder+r'\temp'):
            os.makedirs(folder+r'\temp')
        temp_dir = folder+r'\temp'

        dfs = {}
        if not data_path.endswith('.xlsx'):
            # NEW (convert to XLSX file)
            print('File found in XLS format. Converting and reading again.')
            data_path = self.convert_xls2xlsx(xls_file_path=data_path, target_folder=temp_dir)
            excel_file = pd.ExcelFile(data_path)
            sheet_names = [(id, sheet) for id, sheet in enumerate(excel_file.sheet_names) if sheet.startswith('Fzg')]
            dfs = {sheet: excel_file.parse(sheet, header=None) for id, sheet in sheet_names}
            matches = {sheet: self.find_string_in_dataframe(dfs[sheet]) for id, sheet in sheet_names}

            
        workbook = load_workbook(excel_file, data_only=True)
        for id, sheet_name in sheet_names:
            match_trigger_metadata = matches[sheet_name]
            for row_skip_trigger_metadata, col_skip_trigger_metadata in match_trigger_metadata:
                break

            sheet_select = workbook[sheet_name]
            log_file_name = 'log_folder-'+subfolder+'_file-'+filename+'-'+ext[1:]+'_sheet-'+sheet_name+'.log'
            log_path = os.path.join(folder, log_file_name)
            logger = self.setup_logger(log_path)
            logger.info('Funktion FORMATIERUNG gestartet')
            print(f'Read sheet: {sheet_name}')
            logger.debug(f'Read sheet: {sheet_name}')
            
            # Dataframe for filtered data
            filtered_data = []

            # NEW: use xlsxDecoder class
            xls_object = xlsxDecoder(xlsx_file=data_path)
            cell_data = xls_object.decoder(sheet_nr = id+1, logger = logger)    # +1 since enumerate starts with index = 0
            logger.debug('Start reading cell data according to valid format. Check:')
            logger.debug('1. Is color valid? If YES then do 2. If NO do not consider those values and check the next cell.')
            logger.debug('2. Are the values struck out? If YES consider values. If NO do not consider values.')
            for row in sheet_select.iter_rows():
                filtered_row = []
                for cell in row:
                    logger.debug('--------------------------------------')
                    cell_coordinate = cell.coordinate
                    logger.debug(f'Read cell: {cell.coordinate}')
                
                    if cell_coordinate not in cell_data.keys():
                        logger.debug(f'Cell {cell.coordinate} not in range of active cells in the given row.')
                        continue

                    # Check the cell of valid coloring and struck outs
                    cell_format = cell_data[cell_coordinate]
                    if cell_format['value'] is np.nan or cell_format['value'] is None:
                        filtered_row.append(None)
                    elif len(cell_format['value']) > 1:  # nested values of different styles in selected cell
                        logger.debug(f'Values with different styles found in cell {cell_coordinate}.')
                        logger.debug(f"Values: {cell_format['value']}")
                        cell_valid_values = []
                        for id, sub_value in enumerate(cell_format['value']):
                            logger.debug(f"Investigating Subvalue: {sub_value}")
                            if self.compare_font_color(cell_format['style'][id]['color']):
                                logger.debug('Color is valid?: True')
                                logger.debug(f"Color found: {cell_format['style'][id]['color']}")
                                if not cell_format['style'][id]['strike']:
                                    logger.debug('Struck out?: False')
                                    cell_valid_values.append(sub_value)
                                    if cell.row == row_skip_trigger_metadata + 1 or cell.column == col_skip_trigger_metadata or cell.column == col_skip_trigger_metadata + 1:
                                        logger.debug('Trigger of metadata found! Format of value not of interest!')
                                else:
                                    if cell.row == row_skip_trigger_metadata + 1 or cell.column == col_skip_trigger_metadata or cell.column == col_skip_trigger_metadata + 1:
                                        cell_valid_values.append(sub_value)
                                        logger.debug('Trigger of metadata found! Format of value not of interest!')
                                    else:
                                        logger.debug('Struck out?: True')
                            else:
                                if cell.row == row_skip_trigger_metadata + 1 or cell.column == col_skip_trigger_metadata or cell.column == col_skip_trigger_metadata + 1:
                                    filtered_row.append(cell_format['value'])
                                    logger.debug('Trigger of metadata found! Format of value not of interest!')
                                else:
                                    logger.debug('Color is valid?: False')
                                    logger.debug(f"Color found: {cell_format['style'][id]['color']}")

                        cell_valid_values = ' '.join(cell_valid_values) # join by spaces
                        cell_valid_values = ' '.join(cell_valid_values.split()) # remove double spaces if exist
                        cell_valid_values = cell_valid_values.replace('\n','').replace('\r', '') # remove syntax commands
                        filtered_row.append(cell_valid_values)
                        logger.debug(f"Valid Values in cell: {cell_valid_values}")
                    else:
                        logger.debug(f"Value: {cell_format['value']}")
                        if self.compare_font_color(cell_format['style'][0]['color']):
                            logger.debug('Color is valid?: True')
                            logger.debug("Color found: {cell_format['style'][0]['color']}")
                            if not cell_format['style'][0]['strike']:
                                logger.debug('Struck out?: False')
                                filtered_row.append(cell_format['value'][0])
                                if cell.row == row_skip_trigger_metadata + 1 or cell.column == col_skip_trigger_metadata or cell.column == col_skip_trigger_metadata + 1:
                                    logger.debug('Trigger of metadata found! Format of value not of interest!')
                            else:
                                if cell.row == row_skip_trigger_metadata + 1 or cell.column == col_skip_trigger_metadata or cell.column == col_skip_trigger_metadata + 1:
                                    filtered_row.append(cell_format['value'][0])
                                    logger.debug('Trigger of metadata found! Format of value not of interest!')
                                else:
                                    logger.debug('Struck out?: True')
                        else:
                            if cell.row == row_skip_trigger_metadata + 1 or cell.column == col_skip_trigger_metadata or cell.column == col_skip_trigger_metadata + 1:
                                filtered_row.append(cell_format['value'][0])
                                logger.debug('Trigger of metadata found! Format of value not of interest!')
                            else:
                                logger.debug('Color is valid?: False')
                                logger.debug("Color found: {cell_format['style'][0]['color']}")
                        logger.debug(f"Valid Value in cell: {cell_format['value'][0]}")
                filtered_data.append(filtered_row)


            # DataFrame aus den gefilterten Daten erstellen
            df_proper = pd.DataFrame(filtered_data)
            dfs[sheet_name] = []
            dfs[sheet_name] = df_proper

            logger.info('Funktion erfolgreich abgeschlossen')
            handlers = logger.handlers[:]
            for handler in handlers:
                handler.close()
                logger.removeHandler(handler)

        return dfs

    def get_dataframe_info(self, df):
        '''Get the dataframe information.
        
        Attributes:
        df -> pd.DataFrame
            Dataframe of the xlsx file.'''
        pattern = self.get_pattern()

        regex = re.compile(pattern)
        group_names = list(regex.groupindex.keys())

        # # Leeres DataFrame für die extrahierten Werte
        result_df = pd.DataFrame(index=df.columns.tolist(),columns=group_names)

        # Iteriere über jede Spalte von df
        for col in df.columns:
            # Wende das Regex-Pattern auf jede Zeile der Spalte an
            for idx, value in df[col].items():
                regex_match = re.match(pattern, str(value))
                if regex_match:
                    # Extrahiere die benannten Gruppen und ordne sie den Spalten von result_df zu
                    group_values = regex_match.groupdict()
                    for group_name, group_value in group_values.items():
                        if group_value:  # Nur wenn der Wert nicht None oder leer ist
                            result_df.iloc()
                            result_df.at[col, group_name] = group_value

        return result_df

class ParallelPathFinder():
    '''Class to search for files in a directory and its subdirectories in parallel.'''

    def __init__(self, path, allowed_extensions) -> None:
        '''Initialize the ParallelPathFinder class.
        
        Attributes:
        path -> str
            Path to the directory.
        allowed_extensions -> list
            List of allowed extensions.'''
        self.allowed_extensions = allowed_extensions
        self.path = path

    def search_in_directory(self, directory):
        '''Search in a directory for files.
        
        Attributes:
        directory -> str
            Directory to be searched.'''
        matches = []
        try:
            for entry in os.listdir(directory):
                full_path = os.path.join(directory, entry)
                # Prüfe, ob es eine Datei ist
                if os.path.isfile(full_path):
                    # Extrahiere Dateiname und Dateiendung (in Kleinbuchstaben)
                    name, ext = os.path.splitext(entry)
                    ext = ext.lower()
                    # Datei muss mit "Ausrüstung" beginnen und eine erlaubte Excel-Endung haben
                    if entry.startswith("Ausrüstung") and ext in self.allowed_extensions:
                        matches.append(full_path)
        except PermissionError:
            # Überspringe Verzeichnisse, auf die nicht zugegriffen werden kann
            pass
        return matches

    def find_files(self):
        '''Find files in the directory and its subdirectories.'''
        directories = []
        # Mit os.walk alle Verzeichnisse sammeln
        for root, dirs, files in os.walk(self.path):
            directories.append(root)
        
        found_files = []
        # Verwende ThreadPoolExecutor für parallele I/O-Operationen
        with ThreadPoolExecutor() as executor:
            # Starte für jedes Verzeichnis einen Suchjob
            futures = {executor.submit(self.search_in_directory, directory): directory for directory in directories}
            
            for future in as_completed(futures):
                result = future.result()
                if result:
                    found_files.extend(result)
        
        return found_files